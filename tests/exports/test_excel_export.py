"""Tests for Excel workbook export.

Dev Order: CNC-EXCEL-EXPORT-1

Tests verify:
1. Workbook is generated with expected sheets
2. Values match BidV1 fixtures (Python is source of truth)
3. Tier comparison layout is correct
"""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from business.exports.excel import export_production_estimate_workbook

FIXTURES_DIR = Path(__file__).parent.parent.parent / "fixtures" / "bids"


def load_json(path: Path) -> dict:
    """Load JSON file."""
    with open(path) as f:
        return json.load(f)


@pytest.fixture
def production_bids() -> list[dict]:
    """Load all production bid fixtures."""
    return [
        load_json(FIXTURES_DIR / "eco_loom_production_bid_100_v1.json"),
        load_json(FIXTURES_DIR / "eco_loom_production_bid_250_v1.json"),
        load_json(FIXTURES_DIR / "eco_loom_production_bid_500_v1.json"),
    ]


@pytest.fixture
def workbook_path(production_bids, tmp_path) -> Path:
    """Generate workbook and return path."""
    output = tmp_path / "test_workbook.xlsx"
    export_production_estimate_workbook(production_bids, output_path=output)
    return output


class TestWorkbookGeneration:
    """Verify workbook is generated with expected structure."""

    def test_workbook_is_created(self, workbook_path):
        """Workbook file is created."""
        assert workbook_path.exists()

    def test_workbook_has_expected_sheets(self, workbook_path):
        """Workbook contains all expected sheets."""
        wb = load_workbook(workbook_path)
        expected_sheets = [
            "Summary",
            "Inputs",
            "Cost Breakdown",
            "Risk & Margin",
            "Tier Comparison",
            "Assumptions",
        ]
        assert wb.sheetnames == expected_sheets

    def test_summary_sheet_has_project_info(self, workbook_path):
        """Summary sheet contains project information."""
        wb = load_workbook(workbook_path)
        ws = wb["Summary"]

        assert ws["A1"].value == "Eco-Loom Production Estimate"
        assert ws["B3"].value == "Eco-Loom Scrubber Handle Production"
        assert ws["B4"].value == "ECO-LOOM"


class TestValuesMatchFixtures:
    """Verify workbook values match BidV1 fixtures."""

    @pytest.mark.parametrize(
        "tier,expected_quote,expected_per_unit",
        [
            (100, 864.00, 8.64),
            (250, 2058.00, 8.23),
            (500, 4116.00, 8.23),
        ],
    )
    def test_summary_pricing(self, workbook_path, tier, expected_quote, expected_per_unit):
        """Summary sheet contains correct pricing for each tier."""
        wb = load_workbook(workbook_path)
        ws = wb["Summary"]

        tier_row = {100: 9, 250: 10, 500: 11}[tier]
        assert ws.cell(row=tier_row, column=1).value == tier
        assert ws.cell(row=tier_row, column=2).value == expected_quote
        assert ws.cell(row=tier_row, column=3).value == expected_per_unit

    @pytest.mark.parametrize(
        "tier,expected_material,expected_machine,expected_direct",
        [
            (100, 288.00, 144.00, 432.00),
            (250, 669.00, 360.00, 1029.00),
            (500, 1338.00, 720.00, 2058.00),
        ],
    )
    def test_cost_breakdown_values(
        self, workbook_path, tier, expected_material, expected_machine, expected_direct
    ):
        """Cost Breakdown sheet contains correct costs for each tier."""
        wb = load_workbook(workbook_path)
        ws = wb["Cost Breakdown"]

        tier_col = {100: 2, 250: 3, 500: 4}[tier]

        assert ws.cell(row=4, column=tier_col).value == expected_material
        assert ws.cell(row=5, column=tier_col).value == expected_machine
        assert ws.cell(row=6, column=tier_col).value == expected_direct

    @pytest.mark.parametrize(
        "tier,expected_risked,expected_quote",
        [
            (100, 604.80, 864.00),
            (250, 1440.60, 2058.00),
            (500, 2881.20, 4116.00),
        ],
    )
    def test_risk_margin_values(self, workbook_path, tier, expected_risked, expected_quote):
        """Risk & Margin sheet contains correct risked cost and quote."""
        wb = load_workbook(workbook_path)
        ws = wb["Risk & Margin"]

        tier_row = {100: 19, 250: 20, 500: 21}[tier]

        assert ws.cell(row=tier_row, column=1).value == tier
        assert ws.cell(row=tier_row, column=3).value == expected_risked
        assert ws.cell(row=tier_row, column=4).value == expected_quote


class TestTierComparison:
    """Verify tier comparison layout is correct."""

    def test_tier_comparison_headers(self, workbook_path):
        """Tier Comparison sheet has correct column headers."""
        wb = load_workbook(workbook_path)
        ws = wb["Tier Comparison"]

        assert ws.cell(row=3, column=1).value == "Item"
        assert ws.cell(row=3, column=2).value == "100 units"
        assert ws.cell(row=3, column=3).value == "250 units"
        assert ws.cell(row=3, column=4).value == "500 units"

    def test_tier_comparison_row_labels(self, workbook_path):
        """Tier Comparison sheet has correct row labels."""
        wb = load_workbook(workbook_path)
        ws = wb["Tier Comparison"]

        expected_labels = [
            "Material Cost",
            "Machine Cost",
            "Direct Manufacturing Cost",
            "Risked Cost",
            "Quote Price",
            "Price Per Unit",
        ]

        for row_idx, label in enumerate(expected_labels, 4):
            assert ws.cell(row=row_idx, column=1).value == label

    def test_tier_comparison_all_values(self, workbook_path):
        """Tier Comparison sheet contains all expected values."""
        wb = load_workbook(workbook_path)
        ws = wb["Tier Comparison"]

        expected_values = {
            (4, 2): 288.00,  # 100 unit material
            (4, 3): 669.00,  # 250 unit material
            (4, 4): 1338.00,  # 500 unit material
            (5, 2): 144.00,  # 100 unit machine
            (5, 3): 360.00,  # 250 unit machine
            (5, 4): 720.00,  # 500 unit machine
            (6, 2): 432.00,  # 100 unit direct
            (6, 3): 1029.00,  # 250 unit direct
            (6, 4): 2058.00,  # 500 unit direct
            (7, 2): 604.80,  # 100 unit risked
            (7, 3): 1440.60,  # 250 unit risked
            (7, 4): 2881.20,  # 500 unit risked
            (8, 2): 864.00,  # 100 unit quote
            (8, 3): 2058.00,  # 250 unit quote
            (8, 4): 4116.00,  # 500 unit quote
            (9, 2): 8.64,  # 100 unit per unit
            (9, 3): 8.23,  # 250 unit per unit
            (9, 4): 8.23,  # 500 unit per unit
        }

        for (row, col), expected in expected_values.items():
            actual = ws.cell(row=row, column=col).value
            assert actual == expected, f"Row {row}, Col {col}: expected {expected}, got {actual}"


class TestAssumptionsSheet:
    """Verify assumptions sheet content."""

    def test_assumptions_includes_cnc(self, workbook_path):
        """Assumptions sheet lists CNC machining as included."""
        wb = load_workbook(workbook_path)
        ws = wb["Assumptions"]

        content = " ".join(
            str(ws.cell(row=r, column=1).value or "") for r in range(1, 20)
        )
        assert "CNC machining" in content

    def test_assumptions_excludes_finishing(self, workbook_path):
        """Assumptions sheet lists finishing as excluded."""
        wb = load_workbook(workbook_path)
        ws = wb["Assumptions"]

        content = " ".join(
            str(ws.cell(row=r, column=1).value or "") for r in range(1, 20)
        )
        assert "Finishing labor" in content


class TestDefaultOutputPath:
    """Verify default output path behavior."""

    def test_creates_exports_directory(self, production_bids, tmp_path, monkeypatch):
        """Creates exports/workbooks directory if it doesn't exist."""
        monkeypatch.chdir(tmp_path)
        result = export_production_estimate_workbook(production_bids)

        assert result.exists()
        assert result.parent == Path("exports/workbooks")
        assert result.name == "eco_loom_production_estimate_2026_001.xlsx"
