"""Excel workbook export for production estimates.

Dev Order: CNC-EXCEL-EXPORT-1

Python calculates. Excel explains.

The workbook is an export artifact, not the calculator of record.
All values must match the source BidV1 fixtures.
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_EXPORTS_DIR = Path("exports/workbooks")


def _style_header(ws: Any, row: int, cols: int) -> None:
    """Apply header styling to a row."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")


def _style_currency(ws: Any, row: int, col: int) -> None:
    """Apply currency formatting to a cell."""
    cell = ws.cell(row=row, column=col)
    cell.number_format = '"$"#,##0.00'


def _create_summary_sheet(wb: Workbook, bids: list[dict[str, Any]]) -> None:
    """Create the Summary sheet."""
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Eco-Loom Production Estimate"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Project:"
    ws["B3"] = bids[0]["project_name"]
    ws["A4"] = "Customer:"
    ws["B4"] = bids[0]["customer_name"]
    ws["A5"] = "Status:"
    ws["B5"] = "Production Pricing"

    ws["A7"] = "Quantity Options"
    ws["A7"].font = Font(bold=True)

    headers = ["Quantity", "Quote Price", "Price/Unit"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=8, column=col, value=header)
    _style_header(ws, 8, len(headers))

    for row, bid in enumerate(bids, 9):
        pricing = bid["pricing"]
        ws.cell(row=row, column=1, value=pricing["quantity"])
        ws.cell(row=row, column=2, value=pricing["quote_price"])
        _style_currency(ws, row, 2)
        ws.cell(row=row, column=3, value=pricing["price_per_unit"])
        _style_currency(ws, row, 3)

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15


def _create_inputs_sheet(wb: Workbook, bids: list[dict[str, Any]]) -> None:
    """Create the Inputs sheet."""
    ws = wb.create_sheet("Inputs")

    ws["A1"] = "Job Inputs"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "Material Assumptions"
    ws["A3"].font = Font(bold=True)

    inputs = [
        ("Wood Price", "$17.00 / board foot"),
        ("Wood Waste Allowance", "15%"),
        ("Board Feet per Unit", "0.1302 BF"),
        ("Wood Cost per Unit", "$2.54"),
        ("Velcro Roll Price", "$34.00 / 25-yard roll"),
        ("Velcro Procurement", "Roll-based"),
    ]

    for row, (label, value) in enumerate(inputs, 4):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)

    ws["A11"] = "Machine Assumptions"
    ws["A11"].font = Font(bold=True)

    ws["A12"] = "Machine Rate"
    ws["B12"] = "$72.00 / hour"

    ws["A14"] = "Runtime by Quantity"
    ws["A14"].font = Font(bold=True)

    headers = ["Quantity", "Machine Hours"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=15, column=col, value=header)
    _style_header(ws, 15, len(headers))

    runtimes = [(100, 2.0), (250, 5.0), (500, 10.0)]
    for row, (qty, hours) in enumerate(runtimes, 16):
        ws.cell(row=row, column=1, value=qty)
        ws.cell(row=row, column=2, value=hours)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


def _create_cost_breakdown_sheet(wb: Workbook, bids: list[dict[str, Any]]) -> None:
    """Create the Cost Breakdown sheet."""
    ws = wb.create_sheet("Cost Breakdown")

    ws["A1"] = "Cost Breakdown by Tier"
    ws["A1"].font = Font(bold=True, size=12)

    headers = ["Cost Item", "100 units", "250 units", "500 units"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    _style_header(ws, 3, len(headers))

    for bid in bids:
        col = {100: 2, 250: 3, 500: 4}[bid["pricing"]["quantity"]]

        cost_basis = bid["cost_basis"]

        rows = [
            ("Material Cost", cost_basis["direct_material_cost"]),
            ("Machine Cost", cost_basis["machine_time_cost"]),
            ("Direct Manufacturing Cost", cost_basis["base_manufacturing_cost"]),
        ]

        for row_idx, (label, value) in enumerate(rows, 4):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=col, value=value)
            _style_currency(ws, row_idx, col)

    ws.column_dimensions["A"].width = 30
    for col in range(2, 5):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_risk_margin_sheet(wb: Workbook, bids: list[dict[str, Any]]) -> None:
    """Create the Risk & Margin sheet."""
    ws = wb.create_sheet("Risk & Margin")

    ws["A1"] = "Risk Factors & Margin"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "Risk Factor Structure"
    ws["A3"].font = Font(bold=True)

    factors = [
        ("Tool Wear", "5%"),
        ("Manufacturing Contingency", "10%"),
        ("Business Overhead", "10%"),
        ("Engineering Recovery", "15%"),
        ("Total Risk Factors", "40%"),
    ]

    for row, (label, value) in enumerate(factors, 4):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if label == "Total Risk Factors":
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)

    ws["A10"] = "Target Margin"
    ws["A10"].font = Font(bold=True)
    ws["A11"] = "Target Margin %"
    ws["B11"] = "30%"

    ws["A13"] = "Pricing Formula"
    ws["A13"].font = Font(bold=True)
    ws["A14"] = "Risked Cost = Direct Cost × (1 + 40%)"
    ws["A15"] = "Quote Price = Risked Cost / (1 - 30%)"

    ws["A17"] = "Risked Cost by Tier"
    ws["A17"].font = Font(bold=True)

    headers = ["Quantity", "Direct Cost", "Risked Cost", "Quote Price"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=18, column=col, value=header)
    _style_header(ws, 18, len(headers))

    for row, bid in enumerate(bids, 19):
        pricing = bid["pricing"]
        cost_basis = bid["cost_basis"]

        ws.cell(row=row, column=1, value=pricing["quantity"])
        ws.cell(row=row, column=2, value=cost_basis["base_manufacturing_cost"])
        _style_currency(ws, row, 2)
        ws.cell(row=row, column=3, value=pricing["risked_cost"])
        _style_currency(ws, row, 3)
        ws.cell(row=row, column=4, value=pricing["quote_price"])
        _style_currency(ws, row, 4)

    ws.column_dimensions["A"].width = 30
    for col in range(2, 5):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_tier_comparison_sheet(wb: Workbook, bids: list[dict[str, Any]]) -> None:
    """Create the Tier Comparison sheet."""
    ws = wb.create_sheet("Tier Comparison")

    ws["A1"] = "Production Tier Comparison"
    ws["A1"].font = Font(bold=True, size=12)

    headers = ["Item", "100 units", "250 units", "500 units"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    _style_header(ws, 3, len(headers))

    rows = [
        "Material Cost",
        "Machine Cost",
        "Direct Manufacturing Cost",
        "Risked Cost",
        "Quote Price",
        "Price Per Unit",
    ]

    for row_idx, label in enumerate(rows, 4):
        ws.cell(row=row_idx, column=1, value=label)

    for bid in bids:
        col = {100: 2, 250: 3, 500: 4}[bid["pricing"]["quantity"]]
        pricing = bid["pricing"]
        cost_basis = bid["cost_basis"]

        values = [
            cost_basis["direct_material_cost"],
            cost_basis["machine_time_cost"],
            cost_basis["base_manufacturing_cost"],
            pricing["risked_cost"],
            pricing["quote_price"],
            pricing["price_per_unit"],
        ]

        for row_idx, value in enumerate(values, 4):
            ws.cell(row=row_idx, column=col, value=value)
            _style_currency(ws, row_idx, col)

    ws.column_dimensions["A"].width = 30
    for col in range(2, 5):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_assumptions_sheet(wb: Workbook, bids: list[dict[str, Any]]) -> None:
    """Create the Assumptions sheet."""
    ws = wb.create_sheet("Assumptions")

    ws["A1"] = "Bid Assumptions"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "Included in Quote"
    ws["A3"].font = Font(bold=True)

    included = [
        "CNC machining operations",
        "Shop-supplied wood material",
        "Shop-supplied Velcro material",
    ]

    for row, item in enumerate(included, 4):
        ws.cell(row=row, column=1, value=f"• {item}")

    ws["A8"] = "Excluded from Quote (Pending Requirements)"
    ws["A8"].font = Font(bold=True)

    excluded = [
        "Finishing labor",
        "Packaging materials",
        "Shipping and freight",
        "Sales tax and duties",
    ]

    for row, item in enumerate(excluded, 9):
        ws.cell(row=row, column=1, value=f"• {item}")

    ws["A14"] = "Notes"
    ws["A14"].font = Font(bold=True)

    notes = [
        "Python engine is source of truth for all calculations",
        "This workbook is an export artifact for customer review",
        "Pricing valid while stated assumptions remain unchanged",
    ]

    for row, note in enumerate(notes, 15):
        ws.cell(row=row, column=1, value=f"• {note}")

    ws.column_dimensions["A"].width = 60


def export_production_estimate_workbook(
    bids: list[dict[str, Any]],
    output_path: Path | None = None,
    filename: str = "eco_loom_production_estimate_2026_001.xlsx",
) -> Path:
    """Export a production estimate workbook from BidV1 fixtures.

    Args:
        bids: List of BidV1 dictionaries (100, 250, 500 unit tiers)
        output_path: Output file path (default: exports/workbooks/{filename})
        filename: Output filename if output_path not specified

    Returns:
        Path to the written workbook

    Notes:
        - Python engine is source of truth
        - Workbook values must match BidV1 fixture values
        - Formulas are for transparency, not calculation
    """
    if output_path is None:
        output_path = DEFAULT_EXPORTS_DIR / filename

    output_path.parent.mkdir(parents=True, exist_ok=True)

    bids_sorted = sorted(bids, key=lambda b: b["pricing"]["quantity"])

    wb = Workbook()

    _create_summary_sheet(wb, bids_sorted)
    _create_inputs_sheet(wb, bids_sorted)
    _create_cost_breakdown_sheet(wb, bids_sorted)
    _create_risk_margin_sheet(wb, bids_sorted)
    _create_tier_comparison_sheet(wb, bids_sorted)
    _create_assumptions_sheet(wb, bids_sorted)

    wb.save(output_path)
    return output_path
